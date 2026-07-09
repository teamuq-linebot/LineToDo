#!/usr/bin/env python3
"""Export LINE chat history to Markdown / CSV / XLSX (text messages; media shown
as [image]/[video]/... placeholders since the bytes aren't in the DB).

  export_chat.py --name "Alice" --format xlsx
  export_chat.py --name "Alice" --format md
  export_chat.py --all-csv                  # every message, one CSV, with a chat column
"""
import argparse, csv, os, re
import linedb, linekey

CT = linedb.CT  # contentType -> label


def safe_name(s):
    return re.sub(r"[^\w一-鿿().\-]+", "_", s or "chat")[:48]


def rows_for_chat(con, cid):
    me = linedb.my_mid(con)
    out = []
    for t, frm, text, ct, mid in con.execute(
        "SELECT _createdTime,_from,_text,_contentType,_id FROM _message "
        "WHERE _chatId=? ORDER BY _createdTime", (cid,)
    ).fetchall():
        body = text if text else (CT.get(ct) or f"[type={ct}]")
        out.append({
            "time": linedb.iso(t) or "",
            "direction": "out" if frm == me else "in",
            "sender": "me" if frm == me else (linedb.chat_name(con, frm) or frm or ""),
            "type": "text" if (text and ct == 0) else (CT.get(ct) or f"type{ct}"),
            "text": (body or "").replace("\r\n", "\n"),
        })
    return out


def write_md(path, title, rows):
    L = [f"# LINE — {title}", "", f"> {len(rows)} messages. Media shown as placeholders.", ""]
    cur = None
    for r in rows:
        day = r["time"][:10]
        if day != cur:
            L += ["", f"## {day}", ""]; cur = day
        who = "me" if r["direction"] == "out" else r["sender"]
        body = r["text"].replace("\n", " / ")
        L.append(f"- `{r['time'][11:16]}` **{who}**: {body}")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


def write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["time", "direction", "sender", "type", "text"])
        w.writeheader(); w.writerows(rows)


def write_xlsx(path, title, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook(); ws = wb.active; ws.title = safe_name(title)[:31] or "chat"
    headers = ["time", "direction", "sender", "type", "text"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["time"], r["direction"], r["sender"], r["type"], r["text"]])
    widths = {"A": 20, "B": 10, "C": 18, "D": 10, "E": 90}
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--name")
    ap.add_argument("--format", choices=["md", "csv", "xlsx"], default="xlsx")
    ap.add_argument("--out")
    ap.add_argument("--all-csv", action="store_true",
                    help="export EVERY message across all chats to one CSV")
    a = ap.parse_args()
    con = linedb.open_db()

    if a.all_csv:
        out = a.out or os.path.join(linekey.REPO_ROOT, "export_all.csv")
        chats = linedb.list_chats(con, 100000)
        n = 0
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["chat", "isGroup", "time", "direction", "sender", "type", "text"])
            for ch in chats:
                for r in rows_for_chat(con, ch["chatId"]):
                    w.writerow([ch["name"], ch["isGroup"], r["time"], r["direction"],
                                r["sender"], r["type"], r["text"]]); n += 1
        print(f"[+] wrote {out}  ({n} messages from {len(chats)} chats)")
        return

    if not a.name:
        raise SystemExit("give --name <chat> or --all-csv")
    cid = linedb.resolve_chat(con, a.name)
    title = linedb.chat_name(con, cid) or cid
    rows = rows_for_chat(con, cid)
    out = a.out or os.path.join(linekey.REPO_ROOT, f"export_{safe_name(title)}.{a.format}")
    if a.format == "md":
        write_md(out, title, rows)
    elif a.format == "csv":
        write_csv(out, rows)
    else:
        write_xlsx(out, title, rows)
    print(f"[+] wrote {out}  ({len(rows)} messages, chat='{title}')")


if __name__ == "__main__":
    main()
